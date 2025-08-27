# using xlsxwriter instead of openpyxl, because openpyxl does not support FILTER excel function
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

INPUT_FILE = "input.xlsx"
OUTPUT_FILE = "output.xlsx"
WORKSHEET_TITLE = "seeded_data"

wb_read = load_workbook(INPUT_FILE)
ws_read = wb_read[WORKSHEET_TITLE]

max_col = ws_read.max_column
max_row = ws_read.max_row

data = []
for row in range(1, max_row + 1):
    row_data = []
    for col in range(1, max_col + 1):
        cell_value = ws_read.cell(row=row, column=col).value
        row_data.append(cell_value if cell_value is not None else "")
    data.append(row_data)

wb_read.close()

workbook = xlsxwriter.Workbook(OUTPUT_FILE)

worksheet = workbook.add_worksheet(WORKSHEET_TITLE)

for row_idx, row_data in enumerate(data):
    for col_idx, cell_value in enumerate(row_data):
        worksheet.write(row_idx, col_idx, cell_value)

formula_col = max_col
worksheet.write(0, formula_col, "Formula Result")

FORMULA_CONDITION = "Apple"
UNDEFINED_VALUE = "[0, undefined]"
EXTRA_FEATURE_VALUE = "feature-red"
EXTRA_FEATURE_OBJECT = f"[0, {EXTRA_FEATURE_VALUE}]"

for row in range(2, max_row + 1):
    excel_row = row + 1
    formula = f"""=IF(B{excel_row}="{FORMULA_CONDITION}",
            IF(ISNUMBER(SEARCH("{UNDEFINED_VALUE}", C{excel_row})),
                SUBSTITUTE(C{excel_row}, "{UNDEFINED_VALUE}", "{EXTRA_FEATURE_OBJECT}"),
                IF(ISERROR(SEARCH("{EXTRA_FEATURE_VALUE}", C{excel_row})),
                    C{excel_row}&IF(TRIM(C{excel_row})="", "{EXTRA_FEATURE_OBJECT}", ", [" & (LEN(C{excel_row})-LEN(SUBSTITUTE(C{excel_row},"[",""))) & ", {EXTRA_FEATURE_VALUE}]"),
                    ""
                )
            ),
            ""
        )"""
    worksheet.write(row - 1, formula_col, formula)


worksheet_filtered = workbook.add_worksheet("filtered_data")


worksheet_filtered.write(0, 0, "slug")
worksheet_filtered.write(0, 1, "filter")

formula_res_col = get_column_letter(formula_col + 1)

filter_formula = f"=FILTER('{WORKSHEET_TITLE}'!A2:A{max_row}, '{WORKSHEET_TITLE}'!{formula_res_col}2:{formula_res_col}{max_row}<>\"\")"
worksheet_filtered.write(1, 0, filter_formula)

filter_formula_2 = f"=FILTER('{WORKSHEET_TITLE}'!{formula_res_col}2:{formula_res_col}{max_row}, '{WORKSHEET_TITLE}'!{formula_res_col}2:{formula_res_col}{max_row}<>\"\")"
worksheet_filtered.write(1, 1, filter_formula_2)


workbook.close()
print(f"Script finished! Output saved as {OUTPUT_FILE}")
